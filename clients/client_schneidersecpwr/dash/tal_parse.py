"""Parse a LinkedIn Campaign Manager "Companies" export into the Matched TAL report shape.

WHY A FILE UPLOAD AND NOT AN API CALL. The company-level engagement data behind Plan > Companies is
served by LinkedIn's **Company Intelligence API** (`GET /rest/accountIntelligence`) - whose response
fields are a 1:1 match for this report's columns. That API is documented as private: "available only
to previously approved developers. We are not currently accepting new applications." It is reachable
today only through LinkedIn's certified attribution partners, so no amount of app configuration on
our side unlocks it. The buyer's Campaign-Manager export therefore stays the source, and this module
takes over at the point the manual work actually starts - normalising, summarising and formatting.
If access is ever granted, keep `normalise()` and swap the caller: the API field names are already
accepted as aliases below.

Accepts .csv and .xlsx, and does not assume the header is the first row: Campaign Manager prepends a
few report-metadata lines, and how many varies by export. Headers are matched on a squashed key
(lowercase, alphanumerics only), so "Paid Impressions", "paid impressions" and "paidImpressions" all
land in the same column. Unrecognised columns are reported, never silently dropped.
"""
import csv
import io
import re

# Canonical field -> every header spelling seen from Campaign Manager, this repo's own reference
# exports, and the Company Intelligence API's JSON field names.
ALIASES = {
    "company_name": ["companyname", "company", "accountname", "name"],
    "company_url": ["linkedinurl", "companypageurl", "companyurl", "linkedinpage",
                    "linkedincompanypage", "companylinkedinurl", "pageurl"],
    "company_website": ["companywebsite", "website", "domain"],
    "engagement_level": ["engagementlevel", "engagement", "engagementscore"],
    "organic_impressions": ["organicimpressions", "organicimpr"],
    "organic_engagements": ["organicengagements", "organicengagement"],
    "paid_impressions": ["paidimpressions", "impressions", "paidimpr"],
    "paid_clicks": ["paidclicks", "clicks"],
    "paid_engagements": ["paidengagements", "paidengagement", "engagements"],
    "paid_video_views": ["paidvideoviews", "videoviews"],
    "paid_conversions": ["paidconversions", "conversions"],
    "paid_leads": ["paidleads", "leads"],
    "paid_qualified_leads": ["paidqualifiedleads", "qualifiedleads"],
    "cost_per_qualified_lead": ["costperqualifiedlead", "cpql",
                                "costinlocalcurrencyperpaidqualifiedlead",
                                "costperpaidqualifiedlead"],
    "members_targeted": ["memberstargeted", "members"],
}
LOOKUP = {alias: field for field, aliases in ALIASES.items() for alias in aliases}

NUMERIC_FIELDS = {"organic_impressions", "organic_engagements", "paid_impressions", "paid_clicks",
                  "paid_engagements", "paid_video_views", "paid_conversions", "paid_leads",
                  "paid_qualified_leads", "cost_per_qualified_lead", "members_targeted"}

# LinkedIn returns engagement levels as VERY_HIGH / HIGH / ...; the report prints them in title case.
LEVEL_MAP = {
    "veryhigh": "Very High", "high": "High", "medium": "Medium",
    "low": "Low", "verylow": "Very Low",
}
LEVEL_ORDER = ["Very High", "High", "Medium", "Low", "Very Low"]


class ParseError(ValueError):
    """Raised with a message meant to be shown to the user, not logged and swallowed."""


def _squash(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return v
    s = str(v).strip().replace(",", "").replace("$", "").replace("A$", "").replace("%", "")
    if not s or s in {"-", "--", "N/A", "n/a"}:
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f.is_integer() else f


def _rows_from_csv(data):
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            text = data.decode(encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
        # Sniff the delimiter: LinkedIn exports are comma-separated, but a spreadsheet round-trip in a
        # European locale turns them semicolon-separated and every row would otherwise parse as one cell.
        sample = text[:4096]
        delim = ";" if sample.count(";") > sample.count(",") else ","
        return [r for r in csv.reader(io.StringIO(text), delimiter=delim)]
    raise ParseError("Could not decode that CSV - save it as UTF-8 and try again.")


def _rows_from_xlsx(data):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [["" if c is None else c for c in row] for row in ws.iter_rows(values_only=True)]


def _find_header(rows):
    """Locate the header row. Campaign Manager prefixes exports with report metadata, and the number
    of preamble lines varies, so the header is found by content: the first row within the first 20
    that maps a company-name column AND at least one metric."""
    for i, row in enumerate(rows[:20]):
        mapped = {LOOKUP.get(_squash(c)) for c in row}
        if "company_name" in mapped and len(mapped & NUMERIC_FIELDS) >= 1:
            return i
    for i, row in enumerate(rows[:20]):
        if "company_name" in {LOOKUP.get(_squash(c)) for c in row}:
            return i
    raise ParseError(
        "No company column found. This should be the Campaign Manager export from "
        "Plan > Companies (a column named 'Company Name' or 'Company').")


def normalise(data, filename=""):
    """bytes -> {rows, summary, warnings, columns}. Rows are sorted by paid impressions descending,
    the order the report is specified in."""
    if not data:
        raise ParseError("That file is empty.")
    lower = (filename or "").lower()
    if lower.endswith(".xlsx") or data[:2] == b"PK":
        raw = _rows_from_xlsx(data)
    else:
        raw = _rows_from_csv(data)
    raw = [r for r in raw if any(str(c).strip() for c in r)]
    if not raw:
        raise ParseError("That file has no rows.")

    h = _find_header(raw)
    header = raw[h]
    idx, unknown = {}, []
    for col, cell in enumerate(header):
        field = LOOKUP.get(_squash(cell))
        if field and field not in idx:
            idx[field] = col
        elif str(cell).strip():
            unknown.append(str(cell).strip())

    out = []
    for row in raw[h + 1:]:
        name = str(row[idx["company_name"]]).strip() if idx["company_name"] < len(row) else ""
        if not name:
            continue
        # A trailing "Total"/"Grand total" line would otherwise become a company with enormous
        # metrics and sit at the top of a client-facing table.
        if _squash(name) in {"total", "grandtotal", "totals", "allcompanies"}:
            continue
        rec = {}
        for field, col in idx.items():
            v = row[col] if col < len(row) else None
            rec[field] = _num(v) if field in NUMERIC_FIELDS else (
                str(v).strip() if v not in (None, "") else None)
        rec["engagement_level"] = LEVEL_MAP.get(_squash(rec.get("engagement_level")),
                                                rec.get("engagement_level"))
        out.append(rec)

    if not out:
        raise ParseError("Found the header but no company rows underneath it.")
    out.sort(key=lambda r: (r.get("paid_impressions") or 0), reverse=True)

    warnings = []
    missing = [f for f in ("engagement_level", "paid_impressions") if f not in idx]
    if missing:
        warnings.append("This export has no " + " or ".join(m.replace("_", " ") for m in missing)
                        + " column, so that part of the summary will be blank.")
    if unknown:
        warnings.append("Columns not used: " + ", ".join(unknown[:8])
                        + ("..." if len(unknown) > 8 else ""))

    return {"rows": out, "summary": summarise(out), "warnings": warnings,
            "columns": sorted(idx)}


def summarise(rows):
    """Totals for the Summary sheet. `reached` counts companies with any paid impression - the
    report's "reached by paid ads" line - and is left at 0 if the export carries no such column,
    rather than being inferred from a different metric."""
    total = len(rows)
    by_level = {lvl: 0 for lvl in LEVEL_ORDER}
    other = 0
    for r in rows:
        lvl = r.get("engagement_level")
        if lvl in by_level:
            by_level[lvl] += 1
        elif lvl:
            other += 1
    reached = sum(1 for r in rows if (r.get("paid_impressions") or 0) > 0)
    return {
        "total": total,
        "by_level": by_level,
        "unclassified": other,
        "reached": reached,
        "not_reached": total - reached,
        "paid_impressions": sum(r.get("paid_impressions") or 0 for r in rows),
        "paid_clicks": sum(r.get("paid_clicks") or 0 for r in rows),
        "paid_leads": sum(r.get("paid_leads") or 0 for r in rows),
    }
