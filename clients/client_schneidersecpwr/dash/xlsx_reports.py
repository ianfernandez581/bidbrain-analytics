"""Client-ready Excel builders for the dashboard's Reports tab.

Two reports the media buyer used to compile by hand, rendered to .xlsx server-side:

    build_targeting_xlsx(payload)  ->  "Targeting Breakdown" + "Job Titles Summary"
    build_tal_xlsx(payload)        ->  "Summary" + "Matched Companies"

WHY SERVER-SIDE. The house style below (grey headers, Calibri 9, alternating fills, thousands
separators, frozen header, auto-filter) needs real cell styling. The browser-side spreadsheet
libraries that would run in the dashboard write values only - the free SheetJS build drops fonts and
fills entirely - so a client-facing file built in the browser would come out unformatted. openpyxl in
this Flask service produces the same bytes as the reference files this was matched against.

STYLE IS THE CONTRACT. The constants in the STYLE block were read back out of the two reference
workbooks (`2463_SE_Industrial_Edge_W3_Matched_TAL_Accounts.xlsx`,
`2305_SE_Software_First_Matched_TAL_Accounts.xlsx`): header fill D9D9D9, alternating row fill F7F7F7,
Calibri 9 throughout, bold only on the header row and the title, thin D9D9D9 bottom rule, `#,##0` on
counts, freeze at A2, auto-filter across the table. No colour beyond those greys - the brief is a
plain, printable client document, not a themed one. Change them here and both reports move together.

These builders never invent numbers. Anything the caller did not supply is written as an empty cell,
because a zero in a client report reads as a measured zero.
"""
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- STYLE ------------------------------------------------------------------------------------
FONT_NAME = "Calibri"
SZ = 9
GREY_HEADER = "D9D9D9"
GREY_ALT = "F7F7F7"
GREY_RULE = "D9D9D9"
MUTE_INK = "666666"
NUM_FMT = "#,##0"
MONEY_FMT = '#,##0.00'

F_TITLE = Font(name=FONT_NAME, size=12, bold=True)
F_SUB = Font(name=FONT_NAME, size=SZ, color=MUTE_INK)
F_HEAD = Font(name=FONT_NAME, size=SZ, bold=True)
F_BODY = Font(name=FONT_NAME, size=SZ)

FILL_HEAD = PatternFill("solid", fgColor=GREY_HEADER)
FILL_ALT = PatternFill("solid", fgColor=GREY_ALT)
RULE = Border(bottom=Side(style="thin", color=GREY_RULE))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Engagement levels in the order Campaign Manager reports them (best first). Also the row order of
# the Summary sheet's breakdown - a fixed ladder, so a level with no companies still shows as 0
# rather than disappearing and making the table look shorter than it is.
ENGAGEMENT_LEVELS = ["Very High", "High", "Medium", "Low", "Very Low"]


def _title_block(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = F_TITLE
    ws["A2"] = subtitle
    ws["A2"].font = F_SUB


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = CENTER
        c.border = RULE


def _as_number(v):
    """Return a number for anything that is one, else None.

    Campaign Manager exports arrive with thousands separators, currency symbols and the odd '-' for
    "no value"; a string that merely LOOKS numeric must land in the cell as a number or Excel will
    not sum or sort it. Anything genuinely non-numeric (a range like '50,000-100,000') is left to be
    written as text rather than mangled into a wrong number.
    """
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "").replace("$", "").replace("A$", "")
    if not s or s in {"-", "--", "N/A", "n/a"}:
        return None
    if not re.fullmatch(r"-?\d+(\.\d+)?", s):
        return None
    f = float(s)
    return int(f) if f.is_integer() else f


def _body_cell(ws, row, col, value, *, alt, numeric=False, money=False, wrap=False):
    """One data cell in house style. `numeric` centres and applies the thousands format."""
    n = _as_number(value) if numeric else None
    c = ws.cell(row=row, column=col, value=(n if n is not None else (value if value not in (None, "") else None)))
    c.font = F_BODY
    c.border = RULE
    if alt:
        c.fill = FILL_ALT
    if numeric and n is not None:
        c.number_format = MONEY_FMT if money else NUM_FMT
        c.alignment = CENTER
    elif numeric:
        c.alignment = CENTER          # keep the column visually aligned even when the cell is text
    elif wrap:
        c.alignment = LEFT_WRAP
    return c


def _finish_table(ws, header_row, last_row, last_col):
    """Freeze the header and switch the auto-filter on across the whole table."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if last_row > header_row:
        ws.auto_filter.ref = (f"A{header_row}:"
                              f"{get_column_letter(last_col)}{last_row}")


def _save(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- REPORT 1: TARGETING BREAKDOWN -------------------------------------------------------------
TARGETING_HEADERS = [
    "Phase", "Ad Set", "Geo", "Targeting Method",
    "Include Criteria (titles, seniorities, functions)", "Industries",
    "Company List / TAL", "Exclude Criteria", "Audience Size",
]
TARGETING_WIDTHS = [16, 44, 14, 20, 52, 30, 28, 30, 16]

# Phase display order. Anything unrecognised sorts last rather than being dropped.
PHASE_ORDER = {"Awareness": 0, "Consideration": 1, "Conversion": 2, "Retargeting": 3, "Unspecified": 9}


def _include_criteria(row):
    """Titles / seniorities / functions folded into the one 'Include Criteria' column the report asks
    for, each labelled so a reader can tell which facet a value came from."""
    parts = []
    for label, key in (("Job titles", "job_titles"),
                       ("Seniorities", "job_seniorities"),
                       ("Functions", "job_functions")):
        v = (row.get(key) or "").strip()
        if v:
            parts.append(f"{label}: {v}")
    return "\n".join(parts)


def _sorted_adsets(rows):
    return sorted(rows, key=lambda r: (PHASE_ORDER.get(r.get("phase"), 8),
                                       str(r.get("geo") or ""), str(r.get("adset_name") or "")))


def build_targeting_xlsx(payload):
    """payload: {title, subtitle, rows[]} where each row carries phase / adset_name / geo and the
    seeded audience columns (see clients/client_schneidersecpwr/load_targeting.py)."""
    rows = _sorted_adsets(payload.get("rows") or [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Targeting Breakdown"

    _title_block(ws, payload.get("title") or "Targeting Breakdown",
                 payload.get("subtitle") or "")
    _widths(ws, TARGETING_WIDTHS)
    head = 4
    _header_row(ws, head, TARGETING_HEADERS)

    r = head
    for i, row in enumerate(rows):
        r = head + 1 + i
        alt = i % 2 == 1
        vals = [
            row.get("phase"),
            row.get("adset_name"),
            row.get("geo"),
            row.get("targeting_method"),
            _include_criteria(row),
            row.get("industries"),
            row.get("company_list"),
            row.get("exclusions"),
        ]
        for col, v in enumerate(vals, start=1):
            _body_cell(ws, r, col, v, alt=alt, wrap=col >= 4)
        _body_cell(ws, r, 9, row.get("audience_size"), alt=alt, numeric=True)
    _finish_table(ws, head, r, len(TARGETING_HEADERS))

    _job_titles_sheet(wb, rows)
    return _save(wb)


def _job_titles_sheet(wb, rows):
    """Sheet 2 - every job title used, grouped by funnel phase.

    Deliberately a FLAT table (Phase | Job Title | ...) rather than merged phase blocks: it keeps the
    auto-filter and sorting usable, which is the whole point of a quick-reference sheet.
    """
    ws = wb.create_sheet("Job Titles Summary")
    _title_block(ws, "Job Titles Summary",
                 "Every job title targeted, grouped by funnel phase. One row per title per phase.")
    _widths(ws, [16, 46, 14, 40])
    head = 4
    _header_row(ws, head, ["Phase", "Job Title", "Ad Sets", "Used in"])

    # phase -> title -> [ad set names]. Titles are split on the delimiters a human types into the CSV.
    grouped = {}
    for row in rows:
        titles = [t.strip() for t in re.split(r"[;\n]|(?<!\w),(?!\w)", row.get("job_titles") or "")
                  if t.strip()]
        for t in titles:
            grouped.setdefault(row.get("phase") or "Unspecified", {}).setdefault(t, []).append(
                row.get("adset_name") or "")

    r = head
    i = 0
    for phase in sorted(grouped, key=lambda p: (PHASE_ORDER.get(p, 8), p)):
        for title in sorted(grouped[phase]):
            names = grouped[phase][title]
            r = head + 1 + i
            alt = i % 2 == 1
            _body_cell(ws, r, 1, phase, alt=alt)
            _body_cell(ws, r, 2, title, alt=alt, wrap=True)
            _body_cell(ws, r, 3, len(names), alt=alt, numeric=True)
            _body_cell(ws, r, 4, "; ".join(sorted(set(names))), alt=alt, wrap=True)
            i += 1

    if i == 0:
        # No titles recorded yet: say so in the sheet rather than shipping an empty grid that reads
        # as "this campaign targets no one".
        c = ws.cell(row=head + 1, column=1, value="No job titles recorded yet for these ad sets.")
        c.font = F_BODY
        return
    _finish_table(ws, head, r, 4)


# --- REPORT 2: MATCHED TAL AUDIENCE ------------------------------------------------------------
TAL_HEADERS = [
    "Company Name", "LinkedIn URL", "Engagement Level", "Organic Impressions",
    "Organic Engagements", "Paid Impressions", "Paid Clicks", "Paid Engagements",
    "Paid Video Views", "Paid Conversions", "Paid Leads", "Paid Qualified Leads",
    "Cost per Qualified Lead",
]
TAL_KEYS = [
    "company_name", "company_url", "engagement_level", "organic_impressions",
    "organic_engagements", "paid_impressions", "paid_clicks", "paid_engagements",
    "paid_video_views", "paid_conversions", "paid_leads", "paid_qualified_leads",
    "cost_per_qualified_lead",
]
TAL_WIDTHS = [30, 42, 18, 18, 18, 16, 12, 16, 16, 16, 12, 18, 20]
# Columns 4..13 are counts; the last is money.
TAL_NUMERIC_FROM = 4


def build_tal_xlsx(payload):
    """payload: {title, subtitle, tal_name, rows[], summary{}} - `rows` already normalised by
    tal_parse.normalise(). Rows are written in the order given (the caller sorts by paid impressions
    descending, matching the report's own convention)."""
    rows = payload.get("rows") or []
    summary = payload.get("summary") or {}
    wb = Workbook()

    # -- Summary sheet
    ws = wb.active
    ws.title = "Summary"
    _widths(ws, [32, 16, 14])
    _title_block(ws, payload.get("title") or "Matched TAL Report", payload.get("subtitle") or "")

    ws["A4"] = "Total matched companies"
    ws["A4"].font = F_BODY
    ws["B4"] = summary.get("total") or len(rows)
    ws["B4"].font = F_BODY
    ws["B4"].number_format = NUM_FMT
    if payload.get("tal_name"):
        ws["A5"] = "TAL name"
        ws["A5"].font = F_BODY
        ws["B5"] = payload["tal_name"]
        ws["B5"].font = F_BODY

    _header_row(ws, 7, ["Engagement Level", "Companies", "% of Total"])
    total = summary.get("total") or len(rows) or 0
    by_level = summary.get("by_level") or {}
    for i, level in enumerate(ENGAGEMENT_LEVELS):
        r = 8 + i
        n = by_level.get(level, 0)
        alt = i % 2 == 1
        _body_cell(ws, r, 1, level, alt=alt)
        _body_cell(ws, r, 2, n, alt=alt, numeric=True)
        c = _body_cell(ws, r, 3, (n / total if total else 0), alt=alt)
        c.number_format = "0.0%"
        c.alignment = CENTER

    ws["A14"] = "Paid delivery"
    ws["A14"].font = F_HEAD
    for i, (label, key) in enumerate((("Companies reached by paid ads", "reached"),
                                      ("Companies not yet reached", "not_reached"))):
        r = 15 + i
        ws.cell(row=r, column=1, value=label).font = F_BODY
        c = ws.cell(row=r, column=2, value=summary.get(key, 0))
        c.font = F_BODY
        c.number_format = NUM_FMT

    # -- Matched Companies sheet
    ws2 = wb.create_sheet("Matched Companies")
    _widths(ws2, TAL_WIDTHS)
    _header_row(ws2, 1, TAL_HEADERS)
    r = 1
    for i, row in enumerate(rows):
        r = 2 + i
        alt = i % 2 == 1
        for col, key in enumerate(TAL_KEYS, start=1):
            numeric = col >= TAL_NUMERIC_FROM
            _body_cell(ws2, r, col, row.get(key), alt=alt, numeric=numeric,
                       money=(key == "cost_per_qualified_lead"))
    _finish_table(ws2, 1, r, len(TAL_HEADERS))
    return _save(wb)
